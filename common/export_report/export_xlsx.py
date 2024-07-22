"""
@Project ：iotscan 
@File    ：export_xlsx.py
@IDE     ：PyCharm 
@Author  ：zhizhuo
@Date    ：2023/10/26 14:13 
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io


class ExportXlsx:
    """
    导出xlsx模块
    """

    def __init__(self):
        """
        初始化函数
        """

    @staticmethod
    def get_excel(data):
        """
        将数据写到xlsx文件
        :param data:写入数据
        :return:二进制数据
        """
        illegal_chars = dict.fromkeys(range(0x00, 0x20))
        translator = str.maketrans(illegal_chars)
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['主机', '端口', '端口协议', '扫描类型', 'URL', '协议', 'CMS', '标题', '状态码', '重定向次数',
                   '服务器', '是否CDN', 'CDN IP列表', 'icon_hash', 'ICP备案', '证书', '国家', '省份', '服务商']
        ws.append(headers)
        column_widths = [15, 8, 10, 10, 20, 8, 20, 20, 8, 12, 20, 12, 20, 15, 20, 20, 12, 12, 12]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

        green_fill = PatternFill(start_color="00FF00",
                                 end_color="00FF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = green_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for item in data:
            item['cdn_ip_list'] = ', '.join(item['cdn_ip_list'])
            if not item['cert']:
                item['cert'] = 'Empty'
            row = [str(value) for value in item.values()]
            row = ['' if str(value) == 'None' else value for value in row]
            row = [value.translate(translator) for value in row]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        file_bytes = io.BytesIO()
        wb.save(file_bytes)
        return file_bytes.getvalue()


ExportXlsx = ExportXlsx()
